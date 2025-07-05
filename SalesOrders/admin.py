from django.contrib import admin
from django.urls import path
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
import csv

from .models import SalesOrder, SalesOrderLines
from .forms import SalesOrderImportForm
from django.utils.html import format_html

admin.site.site_header = "Bright Technology Limited Admin Panel"
admin.site.site_title = "Bright Technology Admin"
admin.site.index_title = "Welcome to Bright Technology Admin"


class SalesOrderLinesInline(admin.TabularInline):
    model = SalesOrderLines
    extra = 1


@admin.register(SalesOrder)
class SalesOrderAdmin(admin.ModelAdmin):
    list_display = ("order_reference", "customer", "salesperson", "status_badge", "creation_date", "currency", "total")
    change_list_template = "admin/salesorder_changelist.html"
    inlines = [SalesOrderLinesInline]
    search_fields = ["order_reference", "customer", "salesperson", "status", "creation_date", "currency", "total"]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("import-orders/", self.admin_site.admin_view(self.import_orders), name="SalesOrders_salesorder_import_orders"),
            path("export-orders/", self.admin_site.admin_view(self.export_orders), name="SalesOrders_salesorder_export_orders"),
            path("<path:object_id>/print-pdf/", self.admin_site.admin_view(self.print_pdf_view), name="SalesOrders_salesorder_print_pdf"),
        ]
        return custom_urls + urls

    def import_orders(self, request):
        form = SalesOrderImportForm()

        if request.method == "POST":
            form = SalesOrderImportForm(request.POST, request.FILES)
            if form.is_valid():
                file = form.cleaned_data["file"]

            # Try to load the Excel file safely
            try:
                wb = openpyxl.load_workbook(file, data_only=True)
            except Exception as e:
                self.message_user(request, f"❌ Failed to open Excel file: {e}", level="error")
                return redirect("..")  # ✅ Moved inside the exception only

            # Ensure workbook has at least two sheets
            if len(wb.sheetnames) < 2:
                self.message_user(
                    request,
                    "❌ Your Excel file must have at least two sheets: one for Sales Orders and one for Order Lines.",
                    level="error"
                )
                return redirect("..")

            sales_orders_sheet = wb.worksheets[0]
            order_lines_sheet = wb.worksheets[1]

            # --- Import Sales Orders ---
            sales_order_count = 0
            for i, row in enumerate(sales_orders_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    order_data = {
                        "creation_date": row[0],
                        "customer": row[1],
                        "currency": row[2],
                        "order_reference": row[3],
                        "salesperson": row[4],
                        "status": row[5],
                        "total": row[6],
                        "primary_contact": row[7],
                        "finance_contact": row[8],
                        "delivery_address": row[9],
                        "invoice_address": row[10],
                        "email_address": row[11],
                        "delivery_date": row[12],
                        "delivery_office_location": row[13],
                        "tell_no": row[14],
                        "designation": row[15],
                        "department": row[16],
                        "lpo_confirmation_date": row[17],
                        "lpo_date": row[18],
                        "lpo_number": row[19],
                        "comments": row[20],
                    }
                    SalesOrder.objects.update_or_create(
                        order_reference=order_data["order_reference"], defaults=order_data
                    )
                    sales_order_count += 1
                except Exception as e:
                    self.message_user(request, f"⚠️ Row {i} in Sales Orders sheet skipped: {e}", level="warning")

            # --- Import Sales Order Lines ---
            order_line_count = 0
            for j, row in enumerate(order_lines_sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    order = SalesOrder.objects.get(order_reference=row[0])
                except SalesOrder.DoesNotExist:
                    self.message_user(
                        request,
                        f"⚠️ Row {j} skipped: SalesOrder '{row[0]}' not found.",
                        level="warning"
                    )
                    continue

                try:
                    SalesOrderLines.objects.create(
                        order_reference=order,
                        product=row[1],
                        quantity=row[2],
                        unit_price=row[3],
                        cost=row[4],
                        margin=row[5],
                        margin_percentage=row[6],
                    )
                    order_line_count += 1
                except Exception as e:
                    self.message_user(
                        request,
                        f"⚠️ Row {j} in Order Lines sheet skipped: {e}",
                        level="warning"
                    )

            self.message_user(
                request,
                f"✅ Successfully imported {sales_order_count} Sales Orders and {order_line_count} Order Lines.",
                level="success"
            )
            return redirect("..")

        return render(request, "admin/salesorder_import.html", {"form": form})
    

    def export_orders(self, request):
        from openpyxl import Workbook

        # Create Excel workbook and sheets
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sales Orders"
        ws2 = wb.create_sheet(title="Order Lines")

        # Sheet 1: Sales Orders Header
        ws1.append([
            'Creation Date', 'Customer', 'Currency', 'Order Reference', 'Salesperson',
            'Status', 'Total', 'Primary Contact', 'Finance Contact', 'Delivery Address',
            'Invoice Address', 'Email Address', 'Delivery Date', 'Delivery Office Location',
            'Tell No', 'Designation', 'Department', 'LPO Confirmation Date', 'LPO Date',
            'LPO Number', 'Comments'
        ])

        for order in SalesOrder.objects.all():
            ws1.append([
                order.creation_date, order.customer, order.currency, order.order_reference,
                order.salesperson, order.status, order.total, order.primary_contact,
                order.finance_contact, order.delivery_address, order.invoice_address,
                order.email_address, order.delivery_date, order.delivery_office_location,
                order.tell_no, order.designation, order.department, order.lpo_confirmation_date,
                order.lpo_date, order.lpo_number, order.comments
            ])

        # Sheet 2: Order Lines Header
        ws2.append([
            'Order Reference', 'Product', 'Quantity', 'Unit Price',
            'Cost', 'Margin', 'Margin Percentage'
        ])

        for line in SalesOrderLines.objects.select_related('order_reference').all():
            ws2.append([
                line.order_reference.order_reference, line.product, line.quantity,
                line.unit_price, line.cost, line.margin, line.margin_percentage
            ]) 

        # Set response headers
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = 'attachment; filename="sales_orders_export.xlsx"'

        # Save workbook to response
        wb.save(response)
        return response


    def print_pdf_view(self, request, object_id):
        order = get_object_or_404(SalesOrder, pk=object_id)
        order_lines = order.order_lines.all()

        template = get_template("admin/salesorder_pdf.html")
        html = template.render({"order": order, "order_lines": order_lines})

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="SalesOrder_{order.order_reference}.pdf"'

        pisa.CreatePDF(BytesIO(html.encode("UTF-8")), dest=response)
        return response

    def status_badge(self, obj):
        color = {
            "Quotation": "gray",
            "Sales Order": "purple",
            "Confirmed": "green",
            "Cancelled": "red",
        }.get(obj.status, "black")

        return format_html(
            '<span style="padding:3px 8px; background-color:{}; color:white; border-radius:4px;">{}</span>',
            color,
            obj.status
        )

    status_badge.short_description = "Status"
    status_badge.admin_order_field = "status"


class SalesOrderLinesAdmin(admin.ModelAdmin):
    autocomplete_fields = ["order_reference"]
    search_fields = ["order_reference__order_reference", "order_reference__customer"]


admin.site.register(SalesOrderLines, SalesOrderLinesAdmin)
